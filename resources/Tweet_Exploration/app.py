import tweepy
from openpyxl import Workbook # import openpyxl to interact with excel files

consumer_key = 'vqlRioR8Uho2BBmf5iQqGfksA'
consumer_secret = 'piGigr5NNL1RJvL48n5uYBs58e65jlFHSbK3khQ0zQAEzwlHkz'
access_token = '1050438757994962946-6uP1ZzLUaHbcNa0QN4rFnunYqf74T4'
access_token_secret = 'IVuwltHmj89icmkDbObcU4zYAJ8R83hTsdsMOne3Fmw2g'

auth = tweepy.OAuthHandler(consumer_key, consumer_secret) # twitter consumer key and consumer secret
auth.set_access_token(access_token, access_token_secret) # twitter access token and access_token_secret

api = tweepy.API(auth)
wb = Workbook() # make a workbook
ws = wb.active # grab the active worksheet

categories = ['music', 'entertainment', 'sports', 'arts and design', 'news', 'government', 'politics']
def compare_tweets_by_categories(categories, limit): # limit is the number of tweets to get per category with hash or no hash
    i = 1
    for category in categories:
        if i is not 1: # check if not the first loop, then increment i
            i = 3 + i

        # adds the headers names to the excel sheet
        ws['A{}'.format(i)] = f'{category}-without_hash '
        ws['C{}'.format(i+1)] = 'text'
        ws['D{}'.format(i+1)] = 'quote'
        ws['E{}'.format(i+1)] = 'retweet_count'
        ws['F{}'.format(i+1)] = 'favourite_count'
        ws['G{}'.format(i+1)] = 'a_reply'

        tweet_num = 0 # to identify the number of tweet on excel sheet
        for tweet in tweepy.Cursor(api.search, q=category, count=100).items(limit): # iterate tweets without hash
            tweet_num = tweet_num + 1
            # add the details for each tweet without hash in this category
            ws['B{}'.format(i+2)] = f'tweet-{tweet_num} '
            ws['C{}'.format(i+2)] = tweet.text
            ws['D{}'.format(i+2)] = tweet.is_quote_status
            ws['E{}'.format(i+2)] = tweet.retweet_count   
            ws['F{}'.format(i+2)] = tweet.favorite_count    
            ws['G{}'.format(i+2)] = False if tweet.in_reply_to_status_id == None else True   
            i = 1 + i
        i = 3 + i # increment i to add empty row (spacing) between each group for easy readibility

        # adds the headers names to the excel sheet
        ws['A{}'.format(i)] = f'{category}-hash '
        ws['C{}'.format(i+1)] = 'text'
        ws['D{}'.format(i+1)] = 'quote'
        ws['E{}'.format(i+1)] = 'retweet_count'
        ws['F{}'.format(i+1)] = 'favourite_count'
        ws['G{}'.format(i+1)] = 'a_reply'

        tweet_num = 0 # to identify the number of tweet on excel sheet
        for tweet in tweepy.Cursor(api.search, q='@{}'.format(category), count=100).items(limit): # iterate tweets with hash
            tweet_num = tweet_num + 1
            # add the details for each tweet with hash in this category
            ws['B{}'.format(i+2)] = f'tweet-{tweet_num} '
            ws['C{}'.format(i+2)] = tweet.text
            ws['D{}'.format(i+2)] = tweet.is_quote_status
            ws['E{}'.format(i+2)] = tweet.retweet_count
            ws['F{}'.format(i+2)] = tweet.favorite_count
            ws['G{}'.format(i+2)] = False if tweet.in_reply_to_status_id == None else True
            i = 1 + i
        print(f'processed {category}') # show progress message
    wb.save("sample.xlsx") # save the excel sheet

compare_tweets_by_categories(categories, 5) # calls the method, takes two params (list of categories and number of tweets to get per category)
 